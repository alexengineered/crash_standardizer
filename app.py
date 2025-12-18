"""
Crash Data Standardizer - Main Application
MMUCC dictionaries, and matching engine.
"""

import streamlit as st
import pandas as pd
import io
from datetime import datetime
from pathlib import Path
from openpyxl.styles import PatternFill

from mmucc_loader import MMUCCDictionary
from matching_engine import MatchingEngine, MatchMethod


# === PAGE CONFIG ===
st.set_page_config(
    page_title="Crash Data Standardizer",
    layout="wide"
)

st.markdown("""
<style>
.st-emotion-cache-p38tq {
    font-size: 1.75rem !important;
}
</style>
""", unsafe_allow_html=True)


# === INITIALIZE ===
@st.cache_resource
def load_dictionary():
    return MMUCCDictionary()

@st.cache_resource
def load_engine():
    return MatchingEngine(fuzzy_threshold=80)


dictionary = load_dictionary()
engine = load_engine()


# === HELPER FUNCTIONS ===
def detect_column_type(column_name: str, sample_values: list) -> list[str]:
    """Suggest which MMUCC categories might apply to this column."""
    name_lower = column_name.lower()
    suggestions = []

    # Column name hints
    hints = {
        "manner_of_collision": ["manner", "collision", "crash type", "accident type"],
        "injury_severity": ["injury", "severity", "kabco", "injured"],
        "weather_condition": ["weather", "atmospheric"],
        "light_condition": ["light", "lighting", "illumination"],
        "road_surface_condition": ["surface", "road condition", "pavement"],
        "first_harmful_event": ["harmful", "first event", "event"],
        "contributing_factor_driver": ["contributing", "factor", "cause", "driver action"],
        "distracted_by": ["distract", "distraction"],
        "condition_at_time": ["condition", "driver condition", "impair"],
        "junction_type": ["junction", "intersection", "jct"],
        "vehicle_body_type": ["vehicle type", "body type", "veh type"],
        "trafficway_type": ["trafficway", "roadway type", "road type"],
        "traffic_control": ["traffic control", "control device", "signal", "sign"],
    }

    for category, keywords in hints.items():
        for keyword in keywords:
            if keyword in name_lower:
                suggestions.append(category)
                break

    return suggestions


def is_likely_id_column(column_name: str, sample_values: list) -> bool:
    """Check if column appears to be an ID/key column (not for categorization)."""
    name_lower = column_name.lower()

    # Name-based detection
    id_keywords = ["id", "number", "num", "no", "key", "index", "record", "case", "date", "road", "street", "address", "location", "route", "highway"]
    if any(kw in name_lower for kw in id_keywords):
        return True

    # Value-based detection - mostly numeric or alphanumeric codes
    if sample_values:
        numeric_count = sum(1 for v in sample_values if str(v).replace("-", "").replace(".", "").isdigit())
        if numeric_count / len(sample_values) > 0.8:
            return True

    return False


def row_needs_review(code_val, conf_val, std_val) -> bool:
    """
    Determine if a row needs human review.
    Returns True if: code is 99, label is Unknown/empty, or confidence < 100.
    """
    if code_val == "99":
        return True
    if pd.isna(std_val) or str(std_val).strip() == "":
        return True
    if str(std_val).lower() == "unknown":
        return True
    if 0 < conf_val < 100:
        return True
    return False


def process_dataframe(df: pd.DataFrame, column: str, category: str) -> pd.DataFrame:
    """Process a column and return results with standardized values."""
    results = []

    for idx, value in df[column].items():
        text = str(value) if pd.notna(value) else ""
        result = engine.match(text, category)

        results.append({
            "row": idx,
            "original": value,
            "code": result.code,
            "standardized": result.label,
            "confidence": round(result.confidence, 1),
            "method": result.method.value,
            "needs_review": "Yes" if result.needs_review else "No"
        })

    return pd.DataFrame(results)


def build_fact_dataframe(original_df: pd.DataFrame, results: dict) -> pd.DataFrame:
    """
    Build fact table with standardized columns placed next to their source columns.
    Returns DataFrame with Review_Needed first, then source->standardized pairs.
    """
    # Start with Review_Needed column
    fact_df = pd.DataFrame()
    review_needed = ["NO"] * len(original_df)

    # Process each original column
    for col in original_df.columns:
        # Add original column
        fact_df[col] = original_df[col].values

        # If this column was mapped, add its standardized columns right after
        if col in results:
            result_df = results[col]
            fact_df[f"{col}_Code"] = result_df["code"].values
            fact_df[f"{col}_Standardized"] = result_df["standardized"].values
            fact_df[f"{col}_Confidence"] = result_df["confidence"].values

            # Check each row for review conditions
            for idx in range(len(result_df)):
                code_val = result_df.iloc[idx]["code"]
                conf_val = result_df.iloc[idx]["confidence"]
                std_val = result_df.iloc[idx]["standardized"]

                if row_needs_review(code_val, conf_val, std_val):
                    review_needed[idx] = "YES"

    # Insert Review_Needed as first column
    fact_df.insert(0, "Review_Needed", review_needed)

    return fact_df


def export_to_excel(original_df: pd.DataFrame, results: dict, include_dimension_tables: bool = True) -> bytes:
    """
    Export standardized data to Excel.
    Creates fact table + optional dimension tables for Power BI.
    """
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Build fact table with proper column ordering
        fact_df = build_fact_dataframe(original_df, results)
        fact_df.to_excel(writer, sheet_name="Crash_Data", index=False)

        # Apply color formatting
        ws = writer.sheets["Crash_Data"]
        red_fill = PatternFill(start_color="f7706a", end_color="f7706a", fill_type="solid")
        yellow_fill = PatternFill(start_color="fff192", end_color="fff192", fill_type="solid")

        # Find columns that need formatting
        for col_idx, col_name in enumerate(fact_df.columns, 1):
            if col_name.endswith("_Code") or col_name.endswith("_Standardized") or col_name.endswith("_Confidence"):
                for row_idx in range(2, len(fact_df) + 2):  # Start at 2 to skip header
                    cell = ws.cell(row=row_idx, column=col_idx)
                    df_row = row_idx - 2

                    # Get the base column name
                    base_col = col_name.rsplit("_", 1)[0]
                    code_val = fact_df.loc[df_row, f"{base_col}_Code"]
                    conf_val = fact_df.loc[df_row, f"{base_col}_Confidence"]
                    std_val = fact_df.loc[df_row, f"{base_col}_Standardized"]

                    # Red: Unknown, code 99, or empty
                    if code_val == "99" or str(std_val).lower() == "unknown" or pd.isna(std_val) or str(std_val).strip() == "":
                        cell.fill = red_fill
                    # Yellow: confidence below 100 but above 0
                    elif 0 < conf_val < 100:
                        cell.fill = yellow_fill

        # Dimension tables
        if include_dimension_tables:
            for category in dictionary.categories:
                codes = dictionary.get_codes(category)
                dim_df = pd.DataFrame([
                    {"Code": code, "Label": label}
                    for code, label in codes.items()
                ])
                sheet_name = f"Dim_{category}"[:31]
                dim_df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Review flagged items
        review_rows = []
        for col_name, result_df in results.items():
            flagged = result_df[result_df["needs_review"] == "Yes"]
            for _, row in flagged.iterrows():
                review_rows.append({
                    "Column": col_name,
                    "Row": row["row"],
                    "Original": row["original"],
                    "Suggested_Code": row["code"],
                    "Suggested_Label": row["standardized"],
                    "Confidence": row["confidence"],
                    "Method": row["method"]
                })

        if review_rows:
            review_df = pd.DataFrame(review_rows)
            review_df.to_excel(writer, sheet_name="Needs_Review", index=False)

    return output.getvalue()


# === MAIN UI ===
st.title("Crash Data Standardizer")
st.markdown("*Transform messy crash data into clean, standardized categories for Power BI, Tableau & ArcGIS*")


# Sidebar
with st.sidebar:
    st.header("Settings")

    fuzzy_threshold = st.slider(
        "Fuzzy Match Threshold",
        min_value=60,
        max_value=95,
        value=80,
        help="Minimum similarity score for fuzzy matching (lower = more matches, higher = stricter)"
    )
    engine.fuzzy_threshold = fuzzy_threshold

    include_dimensions = st.checkbox(
        "Include Dimension Tables",
        value=True,
        help="Add lookup tables for Power BI"
    )

    st.divider()
    st.markdown("**Categories Available:**")
    for cat in dictionary.categories:
        st.markdown(f"- {dictionary.get_category_label(cat)}")


# File Upload
st.header("1. Upload CSV/Excel")

uploaded_file = st.file_uploader(
    "Upload crash data file",
    type=["csv", "xlsx", "xls"],
    key="data_upload"
)

# Load data into session state
if uploaded_file:
    if uploaded_file.name.endswith('.csv'):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)
    st.session_state['df'] = df
    st.session_state['source'] = 'file'


# Process data if loaded
if 'df' in st.session_state:
    df = st.session_state['df']

    st.header("2. Map Columns to MMUCC Categories")

    # Show data preview
    with st.expander("Preview Data", expanded=False):
        st.dataframe(df.head(10), use_container_width=True)

    # === TWO-TIER COLUMN MAPPING ===
    column_mappings = {}

    # Categorize columns by confidence level
    auto_matched = []      # High confidence auto-detection
    skipped = []           # ID columns, dates, etc.
    needs_attention = []   # No match or ambiguous

    for column in df.columns:
        sample_values = df[column].dropna().head(5).tolist()
        is_id = is_likely_id_column(column, sample_values)
        suggestions = detect_column_type(column, sample_values)

        col_info = {
            "name": column,
            "samples": sample_values,
            "is_id": is_id,
            "suggestions": suggestions
        }

        if is_id:
            skipped.append(col_info)
        elif suggestions:
            auto_matched.append(col_info)
        else:
            needs_attention.append(col_info)

    # --- TIER 1: Auto-matched columns (compact rows) ---
    if auto_matched:
        st.markdown(f"##### Auto-detected ({len(auto_matched)} columns)")

        for col_info in auto_matched:
            col1, col2, col3 = st.columns([3, 4, 1])

            with col1:
                st.markdown(f"**{col_info['name']}**")
                st.caption(f"{', '.join(str(v)[:15] for v in col_info['samples'][:2])}")

            with col2:
                options = ["Skip"] + dictionary.categories
                default_idx = 0
                if col_info['suggestions'] and col_info['suggestions'][0] in options:
                    default_idx = options.index(col_info['suggestions'][0])

                selected = st.selectbox(
                    "Category",
                    options,
                    index=default_idx,
                    key=f"map_{col_info['name']}",
                    format_func=lambda x: x if x == "Skip" else dictionary.get_category_label(x),
                    label_visibility="collapsed"
                )

                if selected != "Skip":
                    column_mappings[col_info['name']] = selected

            with col3:
                st.markdown("OK")

        st.divider()

    # --- TIER 2: Needs attention (expanded) ---
    if needs_attention:
        st.markdown(f"##### Needs selection ({len(needs_attention)} columns)")
        st.caption("No auto-match found - please select a category or skip. *If no category fits, this is likely an identifier (location, name, ID) - choose Skip.*")

        for col_info in needs_attention:
            with st.container(border=True):
                col1, col2 = st.columns([1, 2])

                with col1:
                    st.markdown(f"**{col_info['name']}**")
                    st.caption(f"Sample: {', '.join(str(v)[:20] for v in col_info['samples'][:3])}")

                with col2:
                    options = ["Skip"] + dictionary.categories

                    selected = st.selectbox(
                        "Select category",
                        options,
                        index=0,
                        key=f"map_{col_info['name']}",
                        format_func=lambda x: x if x == "Skip" else dictionary.get_category_label(x)
                    )

                    if selected != "Skip":
                        column_mappings[col_info['name']] = selected

        st.divider()

    # --- TIER 3: Skipped columns (collapsed) ---
    if skipped:
        with st.expander(f"Skipped ({len(skipped)} columns) - ID/date fields", expanded=False):
            st.caption("These look like ID or date columns. Override if needed.")
            for col_info in skipped:
                col1, col2 = st.columns([3, 5])

                with col1:
                    st.markdown(f"{col_info['name']}")
                    st.caption(f"{', '.join(str(v)[:15] for v in col_info['samples'][:2])}")

                with col2:
                    options = ["Skip"] + dictionary.categories

                    selected = st.selectbox(
                        "Category",
                        options,
                        index=0,  # Default to Skip
                        key=f"map_{col_info['name']}",
                        format_func=lambda x: x if x == "Skip" else dictionary.get_category_label(x),
                        label_visibility="collapsed"
                    )

                    if selected != "Skip":
                        column_mappings[col_info['name']] = selected

    # Summary
    if column_mappings:
        st.success(f"**{len(column_mappings)} columns** ready to process")

    # Process button
    if column_mappings:
        st.header("3. Process & Standardize")

        if st.button("Standardize Data", type="primary"):
            results = {}
            progress = st.progress(0)
            status = st.status("Processing...", expanded=True)

            for i, (column, category) in enumerate(column_mappings.items()):
                status.write(f"Processing: {column} -> {category}")
                results[column] = process_dataframe(df, column, category)
                progress.progress((i + 1) / len(column_mappings))

            st.session_state['results'] = results
            status.update(label="Complete!", state="complete")

    # Show results
    if 'results' in st.session_state:
        results = st.session_state['results']

        st.header("4. Review Results")

        # Stats
        total_rows = len(df)
        for col_name, result_df in results.items():
            matched = len(result_df[result_df['code'] != '99'])
            review = len(result_df[result_df['needs_review'] == "Yes"])

            col1, col2, col3, col4 = st.columns(4)
            col1.metric(f"{col_name}", f"{matched}/{total_rows} matched")
            col2.metric("Regex", len(result_df[result_df['method'] == 'regex']))
            col3.metric("Fuzzy", len(result_df[result_df['method'] == 'fuzzy']))
            col4.metric("Needs Review", review)


        # Editable results
        st.markdown("**Edit results as needed:**")

        tabs = st.tabs(list(results.keys()))

        for tab, (col_name, result_df) in zip(tabs, results.items()):
            with tab:
                category = column_mappings[col_name]
                codes = dictionary.get_codes(category)
                code_options = list(codes.keys())

                edited = st.data_editor(
                    result_df,
                    column_config={
                        "row": st.column_config.NumberColumn("Row", disabled=True),
                        "original": st.column_config.TextColumn("Original", disabled=True),
                        "code": st.column_config.SelectboxColumn("Code", options=code_options),
                        "standardized": st.column_config.TextColumn("Label", disabled=True),
                        "confidence": st.column_config.ProgressColumn("Confidence", min_value=0, max_value=100),
                        "method": st.column_config.TextColumn("Method", disabled=True),
                        "needs_review": st.column_config.TextColumn("Review")
                    },
                    use_container_width=True,
                    hide_index=True
                )

                # Update label when code changes
                edited['standardized'] = edited['code'].map(codes)
                results[col_name] = edited

        st.session_state['results'] = results

        # Export
        st.header("5. Export")

        col1, col2 = st.columns(2)

        with col1:
            excel_data = export_to_excel(df, results, include_dimensions)

            # Generate filename from uploaded file
            if uploaded_file and uploaded_file.name:
                base_name = Path(uploaded_file.name).stem
                excel_filename = f"{base_name}_standardized.xlsx"
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                excel_filename = f"standardized_data_{timestamp}.xlsx"

            st.download_button(
                "Download Excel (Power BI Ready)",
                data=excel_data,
                file_name=excel_filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        with col2:
            # CSV export - use same build function for consistency
            fact_df = build_fact_dataframe(df, results)
            csv_data = fact_df.to_csv(index=False)

            # Generate filename from uploaded file
            if uploaded_file and uploaded_file.name:
                base_name = Path(uploaded_file.name).stem
                csv_filename = f"{base_name}_standardized.csv"
            else:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M")
                csv_filename = f"standardized_data_{timestamp}.csv"

            st.download_button(
                "Download CSV",
                data=csv_data,
                file_name=csv_filename,
                mime="text/csv"
            )

# Footer
st.divider()
st.markdown("""
<small>
<b>Standards:</b> Based on <a href="https://www.nhtsa.gov/mmucc" target="_blank">MMUCC 5th/6th Edition</a> (Model Minimum Uniform Crash Criteria)<br>
<b>Matching:</b> Regex -> Fuzzy (rapidfuzz)
</small>

<small style="margin-top: 2rem; display: block; font-size: 0.8rem;">
Created by AlexEngineered.  I’d love your feedback or suggestions.<br/>
<a href="https://docs.google.com/forms/d/e/1FAIpQLSf6BDS9OtoC_2Ue-porIyxJdClSO7vqDUh11biJrGxl_Q0-wQ/viewform" target="_blank">Send feedback via Google Forms</a>
</small>
""", unsafe_allow_html=True)
